import io
import re
import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from database import get_proyectos, get_cuentas, add_gasto, get_clasificaciones_dict, get_clasificaciones_df, add_clasificacion

class DynamicClasificaciones(dict):
    def keys(self):
        return get_clasificaciones_dict().keys()
    def __getitem__(self, key):
        return get_clasificaciones_dict()[key]
    def __iter__(self):
        return iter(get_clasificaciones_dict())
    def __len__(self):
        return len(get_clasificaciones_dict())
    def items(self):
        return get_clasificaciones_dict().items()
    def __contains__(self, item):
        return item in get_clasificaciones_dict()

CLASIFICACIONES = DynamicClasificaciones()

def _clean_excel_name(text, prefix="R_"):
    clean = re.sub(r'[^a-zA-Z0-9_]', '_', str(text))
    if not clean or (not clean[0].isalpha() and clean[0] != '_'):
        clean = '_' + clean
    return f"{prefix}{clean}"

def generate_excel_template():
    """
    Genera un archivo Excel en memoria con formato y validaciones de datos dependientes (cascada)
    para la clasificación de 3 niveles: Rubro Principal -> Subrubro -> Concepto Detallado.
    Retorna los bytes del archivo generado.
    """
    wb = Workbook()
    
    # 1. Crear las hojas
    ws_gastos = wb.active
    ws_gastos.title = "Gastos"
    ws_lists = wb.create_sheet("Listas_Config")
    
    # 2. Rellenar las listas en Listas_Config
    df_proy = get_proyectos(only_active=True)
    proyectos = ('[' + df_proy['codigo'] + '] ' + df_proy['nombre']).tolist() if not df_proy.empty else ["Sin Proyectos Activos"]
    
    clasifs_dict = get_clasificaciones_dict()
    rubros = list(clasifs_dict.keys())
    
    deducibles = ['Sí', 'No']
    estados_fact = ['Pendiente', 'Facturado']
    metodos_pago = ['Tarjeta de Crédito', 'Transferencia Bancaria', 'Efectivo']
    
    # Escribir listas base en Listas_Config
    headers_config = ["Proyectos", "Rubros Principales", "Deducible", "Estado Fact.", "Método Pago"]
    for col_idx, header in enumerate(headers_config, start=1):
        ws_lists.cell(row=1, column=col_idx, value=header).font = Font(bold=True)
        
    # Escribir proyectos (Col A)
    for idx, p in enumerate(proyectos, start=2):
        ws_lists.cell(row=idx, column=1, value=p)
        
    # Escribir rubros (Col B)
    for idx, r in enumerate(rubros, start=2):
        ws_lists.cell(row=idx, column=2, value=r)
        
    # Escribir deducible (Col C)
    for idx, d in enumerate(deducibles, start=2):
        ws_lists.cell(row=idx, column=3, value=d)
        
    # Escribir estados (Col D)
    for idx, e in enumerate(estados_fact, start=2):
        ws_lists.cell(row=idx, column=4, value=e)
        
    # Escribir metodos de pago (Col E)
    for idx, m in enumerate(metodos_pago, start=2):
        ws_lists.cell(row=idx, column=5, value=m)

    # 3. Construir Rangos Nombrados (Defined Names) para la cascada jerárquica
    cur_col = 10  # Comenzar en Columna J (10) en Listas_Config
    
    for r_name, sub_dict in clasifs_dict.items():
        r_named_key = _clean_excel_name(r_name, prefix="R_")
        sub_keys = list(sub_dict.keys())
        if sub_keys:
            ws_lists.cell(row=1, column=cur_col, value=r_named_key).font = Font(bold=True)
            for idx_s, s_name in enumerate(sub_keys, start=2):
                ws_lists.cell(row=idx_s, column=cur_col, value=s_name)
            
            c_letter = get_column_letter(cur_col)
            end_row = 1 + len(sub_keys)
            def_n = DefinedName(r_named_key, attr_text=f"Listas_Config!${c_letter}$2:${c_letter}${end_row}")
            wb.defined_names.add(def_n)
            cur_col += 1
            
            for s_name, conc_list in sub_dict.items():
                if conc_list:
                    s_named_key = _clean_excel_name(s_name, prefix="S_")
                    ws_lists.cell(row=1, column=cur_col, value=s_named_key).font = Font(bold=True)
                    for idx_c, c_name in enumerate(conc_list, start=2):
                        ws_lists.cell(row=idx_c, column=cur_col, value=c_name)
                    
                    c_let_s = get_column_letter(cur_col)
                    end_r_s = 1 + len(conc_list)
                    def_n_s = DefinedName(s_named_key, attr_text=f"Listas_Config!${c_let_s}$2:${c_let_s}${end_r_s}")
                    wb.defined_names.add(def_n_s)
                    cur_col += 1

    # 4. Diseñar la hoja de Gastos
    headers = [
        "Fecha (AAAA-MM-DD)",
        "Concepto General",
        "Monto Neto (IVA Incluido)",
        "Rubro Principal",
        "Subrubro",
        "Concepto Detallado",
        "Proyecto",
        "Deducible (Sí/No)",
        "Estado Facturación",
        "Método Pago",
        "RFC Proveedor (Opcional)",
        "UUID Fiscal (Opcional)"
    ]
    
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    ws_gastos.row_dimensions[1].height = 28
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws_gastos.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        
    col_widths = {
        'A': 18, 'B': 30, 'C': 25, 'D': 25, 'E': 32, 'F': 32,
        'G': 35, 'H': 18, 'I': 20, 'J': 25, 'K': 22, 'L': 38
    }
    for col, width in col_widths.items():
        ws_gastos.column_dimensions[col].width = width

    # 5. Crear Validaciones de Datos Dependientes (Cascada)
    # Rubro Principal (Col D) -> Nivel 1
    dv_rubro = DataValidation(type="list", formula1=f"=Listas_Config!$B$2:$B${len(rubros)+1}", allow_blank=True)
    dv_rubro.error ='El Rubro seleccionado no es válido'
    dv_rubro.errorTitle = 'Rubro Inválido'
    dv_rubro.prompt = 'Seleccione el Rubro Principal'
    dv_rubro.promptTitle = 'Rubro Principal'

    # Subrubro (Col E) -> Nivel 2 (Depende de Col D)
    formula_sub = '=INDIRECT("R_" & SUBSTITUTE(SUBSTITUTE(SUBSTITUTE(SUBSTITUTE(SUBSTITUTE(SUBSTITUTE(D2, " ", "_"), "&", "_"), "-", "_"), "/", "_"), "+", "_"), ".", "_"))'
    dv_subrubro = DataValidation(type="list", formula1=formula_sub, allow_blank=True)
    dv_subrubro.error ='Seleccione un Subrubro válido para el Rubro Principal seleccionado'
    dv_subrubro.errorTitle = 'Subrubro Inválido'
    dv_subrubro.prompt = 'Seleccione el Subrubro correspondiente al Rubro'
    dv_subrubro.promptTitle = 'Subrubro'

    # Concepto Detallado (Col F) -> Nivel 3 (Depende de Col E)
    formula_conc = '=INDIRECT("S_" & SUBSTITUTE(SUBSTITUTE(SUBSTITUTE(SUBSTITUTE(SUBSTITUTE(SUBSTITUTE(E2, " ", "_"), "&", "_"), "-", "_"), "/", "_"), "+", "_"), ".", "_"))'
    dv_concepto = DataValidation(type="list", formula1=formula_conc, allow_blank=True)
    dv_concepto.error ='Seleccione un Concepto Detallado válido para el Subrubro seleccionado'
    dv_concepto.errorTitle = 'Concepto Inválido'
    dv_concepto.prompt = 'Seleccione el Concepto Detallado'
    dv_concepto.promptTitle = 'Concepto Detallado'

    # Proyecto (Col G)
    dv_proy = DataValidation(type="list", formula1=f"=Listas_Config!$A$2:$A${len(proyectos)+1}", allow_blank=True)
    dv_proy.error ='El proyecto seleccionado no es válido'
    dv_proy.prompt = 'Selecciona un proyecto de la lista'

    # Deducible (Col H)
    dv_deduc = DataValidation(type="list", formula1=f"=Listas_Config!$C$2:$C${len(deducibles)+1}", allow_blank=True)
    dv_deduc.error ='Valor inválido (Sí / No)'
    dv_deduc.prompt = 'Selecciona si es deducible/facturable'
    
    # Estado Facturación (Col I)
    dv_est = DataValidation(type="list", formula1=f"=Listas_Config!$D$2:$D${len(estados_fact)+1}", allow_blank=True)
    dv_est.error ='El estado seleccionado no es válido'
    dv_est.prompt = 'Selecciona si ya está Facturado o Pendiente'
    
    # Método Pago (Col J)
    dv_met = DataValidation(type="list", formula1=f"=Listas_Config!$E$2:$E${len(metodos_pago)+1}", allow_blank=True)
    dv_met.error ='El método de pago no es válido'
    dv_met.prompt = 'Selecciona el método de pago'

    # Agregar validaciones a la hoja de Gastos
    ws_gastos.add_data_validation(dv_rubro)
    ws_gastos.add_data_validation(dv_subrubro)
    ws_gastos.add_data_validation(dv_concepto)
    ws_gastos.add_data_validation(dv_proy)
    ws_gastos.add_data_validation(dv_deduc)
    ws_gastos.add_data_validation(dv_est)
    ws_gastos.add_data_validation(dv_met)

    # Asignar rangos de aplicación de validación (filas 2 a 500)
    dv_rubro.add("D2:D500")
    dv_subrubro.add("E2:E500")
    dv_concepto.add("F2:F500")
    dv_proy.add("G2:G500")
    dv_deduc.add("H2:H500")
    dv_est.add("I2:I500")
    dv_met.add("J2:J500")

    # Guardar en memoria y retornar bytes
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()

def import_excel_expenses(file_bytes):
    """
    Lee un archivo Excel cargado por el usuario, valida la jerarquía de 3 niveles de egresos y lo importa a la base de datos.
    """
    try:
        df = pd.read_excel(io.BytesIO(file_bytes), sheet_name="Gastos")
    except Exception as e:
        return {'success': False, 'imported_count': 0, 'errors': [f"Error al leer el archivo Excel: {str(e)}"]}
        
    df.columns = [c.strip() for c in df.columns]
    
    expected_cols = [
        "Fecha (AAAA-MM-DD)",
        "Concepto General",
        "Monto Neto (IVA Incluido)",
        "Rubro Principal",
        "Subrubro",
        "Concepto Detallado",
        "Proyecto",
        "Deducible (Sí/No)",
        "Estado Facturación",
        "Método Pago",
        "RFC Proveedor (Opcional)",
        "UUID Fiscal (Opcional)"
    ]
    
    missing_cols = [col for col in expected_cols[:10] if col not in df.columns]
    if missing_cols:
        return {
            'success': False,
            'imported_count': 0,
            'errors': [f"Faltan columnas obligatorias en la hoja 'Gastos': {', '.join(missing_cols)}"]
        }

    # Obtener catálogos para mapeos de IDs
    df_projects = get_proyectos()
    project_map = dict(zip('[' + df_projects['codigo'] + '] ' + df_projects['nombre'], df_projects['id']))
    
    df_accounts = get_cuentas()
    account_map = {}
    for _, row in df_accounts.iterrows():
        tipo = row['tipo']
        if tipo not in account_map:
            account_map[tipo] = []
        account_map[tipo].append(row['id'])

    errors = []
    rows_to_insert = []
    
    allowed_deduc = {'Sí', 'No'}
    allowed_est = {'Pendiente', 'Facturado'}
    allowed_met = {'Tarjeta de Crédito', 'Transferencia Bancaria', 'Efectivo'}

    for idx, row in df.iterrows():
        row_num = idx + 2
        
        if row.isna().all():
            continue
            
        fecha_val = row.get("Fecha (AAAA-MM-DD)")
        concepto_val = row.get("Concepto General")
        monto_val = row.get("Monto Neto (IVA Incluido)")
        rubro_val = row.get("Rubro Principal")
        subrubro_val = row.get("Subrubro")
        concepto_det_val = row.get("Concepto Detallado")
        proyecto_val = row.get("Proyecto")
        deduc_val = row.get("Deducible (Sí/No)")
        est_val = row.get("Estado Facturación")
        met_val = row.get("Método Pago")
        
        rfc_val = row.get("RFC Proveedor (Opcional)")
        uuid_val = row.get("UUID Fiscal (Opcional)")
        
        # Validar Campos Vacíos
        if pd.isna(fecha_val) or pd.isna(concepto_val) or pd.isna(monto_val) or pd.isna(rubro_val) or pd.isna(subrubro_val) or pd.isna(concepto_det_val) or pd.isna(proyecto_val) or pd.isna(deduc_val) or pd.isna(est_val) or pd.isna(met_val):
            errors.append(f"Fila {row_num}: Contiene campos obligatorios vacíos.")
            continue
            
        # Validar Fecha
        fecha_str = ""
        if isinstance(fecha_val, (datetime.datetime, datetime.date)):
            fecha_str = fecha_val.strftime('%Y-%m-%d')
        else:
            try:
                date_parsed = pd.to_datetime(str(fecha_val).strip(), format='%Y-%m-%d')
                fecha_str = date_parsed.strftime('%Y-%m-%d')
            except Exception:
                errors.append(f"Fila {row_num}: Formato de fecha inválido (AAAA-MM-DD).")
                continue
                
        # Validar Monto
        try:
            monto_float = float(monto_val)
            if monto_float <= 0:
                errors.append(f"Fila {row_num}: El monto debe ser mayor a cero.")
                continue
        except ValueError:
            errors.append(f"Fila {row_num}: El monto debe ser un número válido.")
            continue
            
        # Validar Deducible, Estado, Método
        deduc_str = str(deduc_val).strip()
        if deduc_str not in allowed_deduc:
            errors.append(f"Fila {row_num}: Deducible '{deduc_str}' no es válido.")
            continue
            
        est_str = str(est_val).strip()
        if est_str not in allowed_est:
            errors.append(f"Fila {row_num}: Estado Facturación '{est_str}' no es válido.")
            continue
            
        met_str = str(met_val).strip()
        if met_str not in allowed_met:
            errors.append(f"Fila {row_num}: Método Pago '{met_str}' no es válido.")
            continue
            
        # Validar Proyecto
        proy_str = str(proyecto_val).strip()
        if proy_str not in project_map:
            errors.append(f"Fila {row_num}: Proyecto '{proy_str}' no existe o está inactivo.")
            continue
        proy_id = project_map[proy_str]
        
        # Validar Jerarquía de 3 Niveles
        rubro_str = str(rubro_val).strip()
        subrubro_str = str(subrubro_val).strip()
        concepto_det_str = str(concepto_det_val).strip()
        
        if rubro_str not in CLASIFICACIONES:
            errors.append(f"Fila {row_num}: El Rubro Principal '{rubro_str}' no es válido.")
            continue
            
        if subrubro_str not in CLASIFICACIONES[rubro_str]:
            errors.append(f"Fila {row_num}: El Subrubro '{subrubro_str}' no pertenece al Rubro '{rubro_str}'.")
            continue
            
        if concepto_det_str not in CLASIFICACIONES[rubro_str][subrubro_str]:
            errors.append(f"Fila {row_num}: El Concepto Detallado '{concepto_det_str}' no pertenece al Subrubro '{subrubro_str}'.")
            continue
            
        rfc_str = str(rfc_val).strip() if not pd.isna(rfc_val) else None
        uuid_str = str(uuid_val).strip() if not pd.isna(uuid_val) else None
        
        # Asignar cuenta asociada
        cuenta_id = None
        if met_str in account_map and len(account_map[met_str]) > 0:
            cuenta_id = account_map[met_str][0]
            
        rows_to_insert.append({
            'fecha': fecha_str,
            'concepto': str(concepto_val).strip(),
            'monto_neto': monto_float,
            'rubro': rubro_str,
            'subrubro': subrubro_str,
            'concepto_detallado': concepto_det_str,
            'proyecto_id': proy_id,
            'deducible': deduc_str,
            'estado_facturacion': est_str,
            'metodo_pago': met_str,
            'cuenta_id': cuenta_id,
            'rfc_proveedor': rfc_str,
            'uuid_fiscal': uuid_str
        })

    if errors:
        return {
            'success': False,
            'imported_count': 0,
            'errors': errors
        }
        
    imported_count = 0
    for r in rows_to_insert:
        success, _ = add_gasto(
            fecha=r['fecha'],
            concepto=r['concepto'],
            monto_neto=r['monto_neto'],
            rubro=r['rubro'],
            subrubro=r['subrubro'],
            concepto_detallado=r['concepto_detallado'],
            proyecto_id=r['proyecto_id'],
            deducible=r['deducible'],
            estado_facturacion=r['estado_facturacion'],
            metodo_pago=r['metodo_pago'],
            cuenta_id=r['cuenta_id'],
            rfc_proveedor=r['rfc_proveedor'],
            uuid_fiscal=r['uuid_fiscal']
        )
        if success:
            imported_count += 1
            
    return {
        'success': True,
        'imported_count': imported_count,
        'errors': []
    }

def export_cashflow_matrix_excel(semanas, row_names, df_values, df_status, saldo_inicial):
    from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Flujo de Caja"
    ws.views.sheetView[0].showGridLines = True
    
    # Fonts and fills
    font_title = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    font_section = Font(name="Calibri", size=11, bold=True, color="000000")
    font_header = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    font_regular = Font(name="Calibri", size=10, color="000000")
    font_total = Font(name="Calibri", size=10, bold=True, color="000000")
    
    fill_header = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid") # Dark Blue
    fill_section = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid") # Very light blue
    fill_green = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid") # Light Green (paid)
    fill_pink = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid") # Light Pink (pending)
    fill_totals = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid") # Light Gray
    
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    double_bottom_border = Border(
        top=Side(style='thin', color='000000'),
        bottom=Side(style='double', color='000000'),
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9')
    )
    
    # 1. Header row
    ws.cell(row=1, column=1, value="Movimientos").font = font_header
    ws.cell(row=1, column=1).fill = fill_header
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="left", vertical="center")
    
    for col_idx, q in enumerate(semanas, start=2):
        cell = ws.cell(row=1, column=col_idx, value=q['label'])
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    current_row = 2
    
    # Section ENTRADAS
    ws.cell(row=current_row, column=1, value="ENTRADAS (INGRESOS)").font = font_section
    ws.cell(row=current_row, column=1).fill = fill_section
    for col_idx in range(2, len(semanas) + 2):
        ws.cell(row=current_row, column=col_idx).fill = fill_section
    current_row += 1
    
    inflow_rows = [r for r in row_names if r.startswith("💼") or "Ingresos" in r]
    outflow_rows = [r for r in row_names if not (r.startswith("💼") or "Ingresos" in r)]
    
    inflow_start_row = current_row
    for rname in inflow_rows:
        ws.cell(row=current_row, column=1, value=rname).font = font_regular
        ws.cell(row=current_row, column=1).border = thin_border
        for col_idx, q in enumerate(semanas, start=2):
            val = df_values.loc[rname, q['label']]
            status = df_status.loc[rname, q['label']]
            cell = ws.cell(row=current_row, column=col_idx, value=val)
            cell.font = font_regular
            cell.number_format = '$#,##0.00'
            cell.border = thin_border
            if val > 0:
                cell.fill = fill_green if status == 'Cobrado' else fill_pink
        current_row += 1
    inflow_end_row = current_row - 1
    
    # TOTAL ENTRADAS row
    total_entradas_row = current_row
    ws.cell(row=current_row, column=1, value="TOTAL ENTRADAS").font = font_total
    ws.cell(row=current_row, column=1).fill = fill_totals
    ws.cell(row=current_row, column=1).border = thin_border
    for col_idx, q in enumerate(semanas, start=2):
        col_letter = get_column_letter(col_idx)
        cell = ws.cell(row=current_row, column=col_idx, value=f"=SUM({col_letter}{inflow_start_row}:{col_letter}{inflow_end_row})")
        cell.font = font_total
        cell.number_format = '$#,##0.00'
        cell.fill = fill_totals
        cell.border = thin_border
    current_row += 1
    
    # Section SALIDAS
    ws.cell(row=current_row, column=1, value="SALIDAS (EGRESOS)").font = font_section
    ws.cell(row=current_row, column=1).fill = fill_section
    for col_idx in range(2, len(semanas) + 2):
        ws.cell(row=current_row, column=col_idx).fill = fill_section
    current_row += 1
    
    outflow_start_row = current_row
    for rname in outflow_rows:
        ws.cell(row=current_row, column=1, value=rname).font = font_regular
        ws.cell(row=current_row, column=1).border = thin_border
        for col_idx, q in enumerate(semanas, start=2):
            val = df_values.loc[rname, q['label']]
            status = df_status.loc[rname, q['label']]
            cell = ws.cell(row=current_row, column=col_idx, value=val)
            cell.font = font_regular
            cell.number_format = '$#,##0.00'
            cell.border = thin_border
            if val > 0:
                cell.fill = fill_green if status == 'Pagado' else fill_pink
        current_row += 1
    outflow_end_row = current_row - 1
    
    # TOTAL SALIDAS row
    total_salidas_row = current_row
    ws.cell(row=current_row, column=1, value="TOTAL SALIDAS").font = font_total
    ws.cell(row=current_row, column=1).fill = fill_totals
    ws.cell(row=current_row, column=1).border = thin_border
    for col_idx in range(2, len(semanas) + 2):
        col_letter = get_column_letter(col_idx)
        cell = ws.cell(row=total_salidas_row, column=col_idx, value=f"=SUM({col_letter}{outflow_start_row}:{col_letter}{outflow_end_row})")
        cell.font = font_total
        cell.number_format = '$#,##0.00'
        cell.fill = fill_totals
        cell.border = thin_border
    current_row += 1
    
    # BALANCE row
    balance_row = current_row
    ws.cell(row=current_row, column=1, value="BALANCE (NETO)").font = font_total
    ws.cell(row=current_row, column=1).border = thin_border
    for col_idx, q in enumerate(semanas, start=2):
        col_letter = get_column_letter(col_idx)
        cell = ws.cell(row=current_row, column=col_idx, value=f"={col_letter}{total_entradas_row}-{col_letter}{total_salidas_row}")
        cell.font = font_total
        cell.number_format = '$#,##0.00'
        cell.border = thin_border
    current_row += 1
    
    # SALDO ACUMULADO row
    saldo_row = current_row
    ws.cell(row=current_row, column=1, value="SALDO ACUMULADO").font = font_total
    ws.cell(row=current_row, column=1).border = double_bottom_border
    for col_idx, q in enumerate(semanas, start=2):
        col_letter = get_column_letter(col_idx)
        if col_idx == 2:
            cell = ws.cell(row=current_row, column=col_idx, value=f"={saldo_inicial}+{col_letter}{balance_row}")
        else:
            prev_col_letter = get_column_letter(col_idx - 1)
            cell = ws.cell(row=current_row, column=col_idx, value=f"={prev_col_letter}{saldo_row}+{col_letter}{balance_row}")
        cell.font = font_total
        cell.number_format = '$#,##0.00'
        cell.border = double_bottom_border
        
    # Auto-adjust columns width
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


# ─────────────────────────────────────────────────────────
# HERRAMIENTAS DE EXCEL PARA CATÁLOGO DE CLASIFICACIONES
# ─────────────────────────────────────────────────────────

def export_clasificaciones_excel():
    """
    Exporta el catálogo actual de clasificaciones de la BD a un archivo Excel
    formateado profesionalmente con estilos corporativos J&D.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Catálogo Clasificaciones"
    ws.views.sheetView[0].showGridLines = True

    fill_header = PatternFill(start_color="434E62", end_color="434E62", fill_type="solid")
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_data = Font(name="Calibri", size=10)
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    fill_zebra = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    headers = ["ID", "Rubro Principal", "Subrubro", "Concepto Detallado"]
    ws.append(headers)
    ws.row_dimensions[1].height = 26

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = align_center
        cell.border = thin_border

    df_c = get_clasificaciones_df()
    if not df_c.empty:
        for r_idx, row in enumerate(df_c.itertuples(), start=2):
            ws.append([row.id, row.rubro, row.subrubro, row.concepto])
            ws.row_dimensions[r_idx].height = 20
            
            fill_row = fill_zebra if r_idx % 2 == 0 else PatternFill(fill_type=None)
            for c_idx in range(1, 5):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.font = font_data
                cell.border = thin_border
                if c_idx == 1:
                    cell.alignment = align_center
                else:
                    cell.alignment = align_left
                if r_idx % 2 == 0:
                    cell.fill = fill_row

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 15)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def generate_clasificaciones_template():
    """
    Genera una plantilla vacía con formato corporativo J&D y filas de ejemplo
    para que los usuarios agreguen clasificaciones en lote.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Cargar Clasificaciones"
    ws.views.sheetView[0].showGridLines = True

    fill_header = PatternFill(start_color="434E62", end_color="434E62", fill_type="solid")
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_data = Font(name="Calibri", size=10)
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    headers = ["Rubro Principal", "Subrubro", "Concepto Detallado"]
    ws.append(headers)
    ws.row_dimensions[1].height = 26

    for col_idx in range(1, 4):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = align_center
        cell.border = thin_border

    ejemplos = [
        ("PROYECTOS", "GASTOS", "COMPRA DE HERRAMIENTAS MENORES"),
        ("FIJOS", "Servicios Taller", "MANTENIMIENTO DE AIRE ACONDICIONADO"),
        ("NOMINA", "Sueldos y Salarios", "BONO DE PRODUCTIVIDAD PROYECTO"),
    ]

    for r_idx, (r, s, c) in enumerate(ejemplos, start=2):
        ws.append([r, s, c])
        ws.row_dimensions[r_idx].height = 20
        for c_idx in range(1, 4):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.font = font_data
            cell.alignment = align_left
            cell.border = thin_border

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 40

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def import_clasificaciones_excel(file_bytes):
    """
    Lee un archivo Excel con columnas 'Rubro Principal', 'Subrubro', 'Concepto Detallado'
    e inserta cada clasificación en la BD.
    Retorna (added_count, dup_count, errors)
    """
    try:
        df = pd.read_excel(io.BytesIO(file_bytes))
    except Exception as e:
        return 0, 0, [f"Error al leer el archivo Excel: {str(e)}"]

    col_map = {}
    for col in df.columns:
        c_clean = str(col).strip().upper()
        if 'SUB' in c_clean:
            col_map[col] = 'subrubro'
        elif 'RUBRO' in c_clean:
            col_map[col] = 'rubro'
        elif 'CONCEPTO' in c_clean or 'DETALLE' in c_clean:
            col_map[col] = 'concepto'

    df.rename(columns=col_map, inplace=True)

    required = ['rubro', 'subrubro', 'concepto']
    missing = [r for r in required if r not in df.columns]
    if missing:
        return 0, 0, [f"Columnas faltantes en el archivo Excel: {', '.join(missing)}. El archivo debe contener las columnas 'Rubro Principal', 'Subrubro' y 'Concepto Detallado'."]

    added_count = 0
    dup_count = 0
    errors = []

    for idx, row in df.iterrows():
        rubro = str(row.get('rubro', '')).strip()
        subrubro = str(row.get('subrubro', '')).strip()
        concepto = str(row.get('concepto', '')).strip()

        if not rubro or rubro.lower() == 'nan' or not subrubro or subrubro.lower() == 'nan' or not concepto or concepto.lower() == 'nan':
            continue

        success, msg = add_clasificacion(rubro, subrubro, concepto)
        if success:
            added_count += 1
        else:
            if "existe" in msg.lower() or "unique" in msg.lower():
                dup_count += 1
            else:
                errors.append(f"Fila {idx+2}: {msg}")

    return added_count, dup_count, errors

